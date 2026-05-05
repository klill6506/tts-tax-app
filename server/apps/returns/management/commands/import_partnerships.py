"""
Import partnerships from a Lacerte XLSX export.

Creates the chain: Client -> Entity(partnership) -> TaxYear -> TaxReturn(1065)
plus a backfill of empty FormFieldValues for every line on the 1065.

Default behavior is dry-run + transaction-rolled-back, matching the Lacerte
client-list importer. Pass --commit to actually persist.

Usage:
    # Dry run (default — nothing written)
    poetry run python manage.py import_partnerships \\
        --xlsx-file "D:/tax-test-data/import-sources/TTS Partnerships.xlsx"

    # Actually write
    poetry run python manage.py import_partnerships \\
        --xlsx-file "..." --commit
"""
import datetime
import os
from pathlib import Path

import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from apps.clients.models import Client, Entity, EntityType, TaxYear
from apps.firms.models import Firm
from apps.returns.models import FormDefinition, FormFieldValue, FormLine, TaxReturn


TAX_YEAR = 2025


def normalize_fein(fein):
    if not fein:
        return ""
    digits = "".join(c for c in str(fein) if c.isdigit())
    if len(digits) == 9:
        return f"{digits[:2]}-{digits[2:]}"
    return str(fein).strip()


def normalize_zip(z):
    if z is None:
        return ""
    z = str(int(z)) if isinstance(z, (int, float)) else str(z).strip()
    return z.zfill(5)


class Command(BaseCommand):
    help = (
        "Import partnerships from a Lacerte XLSX export. "
        "Dry-run by default — pass --commit to actually write."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "--xlsx-file",
            type=str,
            required=True,
            help="Absolute path to the Lacerte partnerships XLSX export.",
        )
        parser.add_argument(
            "--commit",
            action="store_true",
            help="Actually write to the DB. Without this, runs in dry-run.",
        )

    def handle(self, *args, **options):
        xlsx_path = Path(options["xlsx_file"])
        commit = options["commit"]

        if not xlsx_path.is_file():
            raise CommandError(f"XLSX not found: {xlsx_path}")

        firm = Firm.objects.filter(name__icontains="Tax Shelter").first()
        if not firm:
            raise CommandError("Firm 'The Tax Shelter' not found.")

        form_def = FormDefinition.objects.filter(code="1065", tax_year_applicable=TAX_YEAR).first()
        if not form_def:
            raise CommandError(
                f"FormDefinition for 1065/{TAX_YEAR} not found. "
                "Run: poetry run python manage.py seed_1065"
            )

        mode = "COMMIT" if commit else "DRY-RUN (no writes)"
        self.stdout.write(f"XLSX: {xlsx_path}")
        self.stdout.write(f"Firm: {firm.name} ({firm.id})")
        self.stdout.write(f"Tax year: {TAX_YEAR}")
        self.stdout.write(f"Mode: {mode}\n")

        wb = openpyxl.load_workbook(str(xlsx_path))
        ws = wb.active
        created = 0
        skipped = 0

        with transaction.atomic():
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
                entity_name = (row[0] or "").strip()
                street = (row[1] or "").strip()
                city = (row[2] or "").strip()
                state = (row[3] or "").strip()
                zip_code = normalize_zip(row[4])
                fein = normalize_fein(row[5])
                date_began = row[6]
                email = (row[8] or "").strip()
                bus_code = str(int(row[9])) if row[9] else ""
                acct_method = (row[10] or "").strip().lower()
                phone = str(row[11] or "").strip()
                principal_activity = (row[13] or "").strip()
                principal_product = (row[14] or "").strip()

                if not entity_name:
                    continue

                self.stdout.write(f"\n--- {entity_name} ({fein}) ---")

                # 1. Client
                client, c_new = Client.objects.get_or_create(firm=firm, name=entity_name)
                self.stdout.write(f"  {'Created' if c_new else 'Exists'} Client: {entity_name}")

                # 2. Entity
                entity = Entity.objects.filter(client=client, entity_type=EntityType.PARTNERSHIP).first()
                if not entity:
                    dt = date_began.date() if isinstance(date_began, datetime.datetime) else date_began
                    entity = Entity.objects.create(
                        client=client,
                        name=entity_name,
                        entity_type=EntityType.PARTNERSHIP,
                        ein=fein,
                        address_line1=street,
                        city=city,
                        state=state,
                        zip_code=zip_code,
                        phone=phone,
                        email=email,
                        naics_code=bus_code,
                        business_activity=principal_activity,
                        date_incorporated=dt,
                    )
                    self.stdout.write(f"  Created Entity: {entity_name}")
                else:
                    self.stdout.write(f"  Entity exists: {entity.name}")

                # 3. TaxYear
                tax_year, ty_new = TaxYear.objects.get_or_create(
                    entity=entity, year=TAX_YEAR,
                    defaults={"filing_states": ["GA"]},
                )
                self.stdout.write(f"  {'Created' if ty_new else 'Exists'} TaxYear: {TAX_YEAR}")

                # 4. TaxReturn
                if TaxReturn.objects.filter(tax_year=tax_year, federal_return__isnull=True).exists():
                    self.stdout.write(f"  Return already exists — skipping")
                    skipped += 1
                    continue

                acct_choice = "accrual" if acct_method == "accrual" else "cash"
                tax_return = TaxReturn.objects.create(
                    tax_year=tax_year,
                    form_definition=form_def,
                    accounting_method=acct_choice,
                    tax_year_start=datetime.date(TAX_YEAR, 1, 1),
                    tax_year_end=datetime.date(TAX_YEAR, 12, 31),
                )
                self.stdout.write(f"  Created TaxReturn: 1065 (id={tax_return.id})")

                # 5. Backfill form field values
                all_lines = FormLine.objects.filter(section__form=form_def)
                FormFieldValue.objects.bulk_create([
                    FormFieldValue(tax_return=tax_return, form_line=ln, value="")
                    for ln in all_lines
                ])
                self.stdout.write(f"  Backfilled {all_lines.count()} form field values")

                # 6. Extra fields
                if bus_code and hasattr(tax_return, "business_activity_code"):
                    tax_return.business_activity_code = bus_code
                if principal_activity and hasattr(tax_return, "product_or_service"):
                    tax_return.product_or_service = principal_activity
                tax_return.save()

                created += 1

            if not commit:
                transaction.set_rollback(True)

        self.stdout.write(f"\n{'='*60}")
        verb = "Created" if commit else "Would create"
        self.stdout.write(self.style.SUCCESS(f"Done. {verb}: {created}, Skipped: {skipped}"))
        if not commit:
            self.stdout.write(self.style.WARNING(
                "DRY-RUN — no changes written. Re-run with --commit to apply."
            ))
